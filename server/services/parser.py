import os
import csv
import io
from typing import List


def extract_text_from_file(filepath: str, filename: str) -> str:
    ext = filename.rsplit(".", 1)[1].lower() if "." in filename else ""

    try:
        if ext == "txt":
            with open(filepath, "r", encoding="utf-8") as f:
                return f.read()

        elif ext == "pdf":
            from PyPDF2 import PdfReader
            reader = PdfReader(filepath)
            text_parts = []
            for page in reader.pages:
                page_text = page.extract_text()
                if page_text:
                    text_parts.append(page_text)
            return "\n".join(text_parts)

        elif ext == "docx":
            from docx import Document
            doc = Document(filepath)
            return "\n".join([para.text for para in doc.paragraphs if para.text.strip()])

        elif ext == "csv":
            with open(filepath, "r", encoding="utf-8") as f:
                return f.read()

        elif ext == "xlsx":
            from openpyxl import load_workbook
            wb = load_workbook(filepath)
            text_parts = []
            for sheet in wb.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    row_text = " | ".join([str(cell) for cell in row if cell is not None])
                    if row_text.strip():
                        text_parts.append(row_text)
            return "\n".join(text_parts)

        else:
            with open(filepath, "r", encoding="utf-8") as f:
                return f.read()

    except Exception as e:
        raise ValueError(f"Failed to extract text from {filename}: {str(e)}")


def parse_questions(filepath: str, ext: str) -> List[str]:
    questions = []

    try:
        if ext == "csv":
            with open(filepath, "r", encoding="utf-8") as f:
                reader = csv.reader(f)
                header = next(reader, None)

                question_col = 0
                if header:
                    for i, col in enumerate(header):
                        col_lower = col.strip().lower()
                        if col_lower in ("question", "questions", "query", "item", "q"):
                            question_col = i
                            break

                for row in reader:
                    if row and len(row) > question_col:
                        question = row[question_col].strip()
                        if question and question != "":
                            questions.append(question)

        elif ext == "xlsx":
            from openpyxl import load_workbook
            wb = load_workbook(filepath)
            ws = wb.active

            question_col = 0
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if header_row:
                for i, cell in enumerate(header_row):
                    if cell and str(cell).strip().lower() in ("question", "questions", "query", "item", "q"):
                        question_col = i
                        break

            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and len(row) > question_col and row[question_col]:
                    question = str(row[question_col]).strip()
                    if question:
                        questions.append(question)

        elif ext in ("pdf", "docx", "txt"):
            text = extract_text_from_file(filepath, f"file.{ext}")
            lines = text.split("\n")

            for line in lines:
                line = line.strip()
                if not line:
                    continue

                if len(line) < 10 and not line.endswith("?"):
                    continue

                if line.endswith("?"):
                    cleaned = _clean_question_number(line)
                    questions.append(cleaned)
                    continue

                if _looks_like_question(line):
                    cleaned = _clean_question_number(line)
                    questions.append(cleaned)

        else:
            raise ValueError(f"Unsupported questionnaire format: {ext}")

    except Exception as e:
        raise ValueError(f"Failed to parse questionnaire: {str(e)}")

    return questions


def _clean_question_number(text: str) -> str:
    import re
    cleaned = re.sub(r"^(?:Q?\d+[\.\)\:\-]|#\d+[\.\)\:\-]?)\s*", "", text, flags=re.IGNORECASE)
    return cleaned.strip() if cleaned.strip() else text.strip()


def _looks_like_question(line: str) -> bool:
    import re
    if re.match(r"^(?:\d+[\.\)\:\-]|Q\d+|[\-\*\•])\s+", line, re.IGNORECASE):
        return True
    return False